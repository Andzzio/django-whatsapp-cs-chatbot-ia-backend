from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.conf import settings
from django.core.cache import cache
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from .models import ProductEmbedding
import logging

logger = logging.getLogger(__name__)


@csrf_exempt
def export_products_excel(request):
    """
    GET /api/products/export/excel/
    Exporta todos los productos a un archivo Excel
    """
    token = request.headers.get("Authorization")
    if token != settings.DASH_TOKEN:
        return HttpResponse("Unauthorized", status=403)

    if request.method != "GET":
        return HttpResponse("Method not allowed", status=405)

    try:
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventario"

        # Headers
        headers = [
            "ID Producto",
            "Nombre",
            "Precio (S/)",
            "Stock S",
            "Stock M",
            "Stock L",
            "Stock XL",
            "Disponible",
            "Categoría",
        ]
        ws.append(headers)

        # Estilos header
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Datos
        products = ProductEmbedding.objects.all().order_by("product_name")
        for product in products:
            ws.append(
                [
                    product.retailer_id,
                    product.product_name,
                    float(product.price) if product.price else 0.0,
                    product.stock_s,
                    product.stock_m,
                    product.stock_l,
                    product.stock_xl,
                    "Sí" if product.is_available else "No",
                    product.category or "",
                ]
            )

        # Ajustar anchos de columna
        ws.column_dimensions["A"].width = 25  # ID
        ws.column_dimensions["B"].width = 45  # Nombre
        ws.column_dimensions["C"].width = 15  # Precio
        ws.column_dimensions["D"].width = 10  # Stock S
        ws.column_dimensions["E"].width = 10  # Stock M
        ws.column_dimensions["F"].width = 10  # Stock L
        ws.column_dimensions["G"].width = 10  # Stock XL
        ws.column_dimensions["H"].width = 15  # Disponible
        ws.column_dimensions["I"].width = 20  # Categoría5

        # Freeze header row
        ws.freeze_panes = "A2"

        # Crear response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="inventario.xlsx"'
        wb.save(response)

        logger.info(f"Excel exported: {products.count()} products")
        return response

    except Exception as e:
        logger.error(f"Error exporting Excel: {e}")
        return HttpResponse(f"Error: {str(e)}", status=500)


@csrf_exempt
def import_products_excel(request):
    """
    POST /api/products/import/excel/
    Importa stock desde un archivo Excel
    Body: multipart/form-data con archivo 'file'
    """
    token = request.headers.get("Authorization")
    if token != settings.DASH_TOKEN:
        return JsonResponse({"error": "Unauthorized"}, status=403)

    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    try:
        excel_file = request.FILES.get("file")
        if not excel_file:
            return JsonResponse({"error": "No file provided"}, status=400)

        # Leer Excel
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb.active

        updated_count = 0
        errors = []
        skipped_count = 0

        # Saltar header (row 1)
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=2, values_only=True), start=2
        ):
            # row = (retailer_id, name, price, stock_s, stock_m, stock_l, stock_xl, available, category)
            if not row or not row[0]:  # Saltar filas vacías o sin ID
                skipped_count += 1
                continue

            retailer_id = str(row[0]).strip()
            stock_s = int(row[3]) if row[3] else 0
            stock_m = int(row[4]) if row[4] else 0
            stock_l = int(row[5]) if row[5] else 0
            stock_xl = int(row[6]) if row[6] else 0
            is_available_text = str(row[7]).strip().lower() if row[7] else "sí"

            try:
                product = ProductEmbedding.objects.get(retailer_id=retailer_id)

                # Actualizar stocks por tallas
                product.stock_s = stock_s
                product.stock_m = stock_m
                product.stock_l = stock_l
                product.stock_xl = stock_xl

                # Actualizar disponibilidad
                if is_available_text in ["sí", "si", "yes", "true", "1"]:
                    product.is_available = True
                elif is_available_text in ["no", "false", "0"]:
                    product.is_available = False
                # Auto-deshabilitar si no hay stock en ninguna talla
                elif product.total_stock == 0:
                    product.is_available = False

                product.save()
                updated_count += 1

            except ProductEmbedding.DoesNotExist:
                errors.append(f"Fila {row_idx}: Producto '{retailer_id}' no encontrado")
            except Exception as e:
                errors.append(f"Fila {row_idx}: {str(e)}")

        # Invalidar cache
        if updated_count > 0 and hasattr(settings, "CATALOG_ID"):
            cache_key = f"catalog_products_{settings.CATALOG_ID}"
            cache.delete(cache_key)

        logger.info(
            f"Excel imported: {updated_count} products updated, "
            f"{skipped_count} skipped, {len(errors)} errors"
        )

        return JsonResponse(
            {
                "status": "success",
                "updated": updated_count,
                "skipped": skipped_count,
                "errors": errors[:10],  # Limitar a 10 errores para no saturar
                "total_errors": len(errors),
            }
        )

    except Exception as e:
        logger.error(f"Error importing Excel: {e}")
        return JsonResponse({"error": str(e)}, status=500)
